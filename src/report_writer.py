"""
Module 6: Report Generation
Generates the query_6.xlsx file with the fixed 22-column schema.
Color-codes rows: green (added), red (removed), yellow (changed cells).
"""

import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# query_6 column schema (22 columns)
QUERY6_COLUMNS = [
    "SAPID",
    "Amendment #",
    "Contract #",
    "Company",
    "Program Name",
    "Product",
    "Formulary Status / BoB",
    "Tier Benefit / Base Rebate %",
    "Price Increase Threshold %",
    "Program Start Date",
    "Program End Date",
    "Payment Terms (Days)",
    "Admin Fee %",
    "Frequency",
    "Commencement Date",
    "Termination Date",
    "Notice Period (Days)",
    "Minimum Rebate",
    "Maximum Rebate",
    "Market Share Requirement",
    "Definition Hash",
    "Notes",
]

# Highlight colors
FILL_ADDED = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
FILL_REMOVED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
FILL_CHANGED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_NORMAL = Font(name="Calibri", size=11)
FONT_REMOVED = Font(name="Calibri", size=11, strikethrough=True, color="C00000")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Map from snapshot product fields to query_6 column names
SNAPSHOT_TO_QUERY6 = {
    "base_rebate_pct": "Tier Benefit / Base Rebate %",
    "formulary_status": "Formulary Status / BoB",
    "price_protection_threshold": "Price Increase Threshold %",
    "start_date": "Program Start Date",
    "end_date": "Program End Date",
    "admin_fee": "Admin Fee %",
    "frequency": "Frequency",
    "minimum_rebate": "Minimum Rebate",
    "maximum_rebate": "Maximum Rebate",
    "market_share": "Market Share Requirement",
    "definition_hash": "Definition Hash",
}


def generate_report(
    delta: dict,
    new_snapshot: dict,
    output_path: str,
) -> str:
    """Generate the query_6.xlsx report from a delta result and snapshot.

    Args:
        delta: Delta report from delta_engine.compare_snapshots().
        new_snapshot: The current snapshot.
        output_path: Path to write the .xlsx file.

    Returns:
        The output file path.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "query_6"

    # Write header row
    _write_header(ws)

    # Write data rows
    row_num = 2
    for change in delta.get("product_changes", []):
        change_type = change["change_type"]
        product_data = change["fields"]
        changed_fields = change.get("changed_fields", {})
        notes = change.get("notes", "")

        # Add agreement-level change notes
        agreement_notes = _format_agreement_changes(delta.get("agreement_changes", []))
        if agreement_notes and change_type != "REMOVED":
            notes = f"{notes}; {agreement_notes}" if notes else agreement_notes

        # Build row values
        row_values = _build_row(product_data, new_snapshot, notes)

        # Write row
        for col_num, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = FONT_NORMAL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Apply row-level highlighting
            if change_type == "ADDED":
                cell.fill = FILL_ADDED
            elif change_type == "REMOVED":
                cell.fill = FILL_REMOVED
                cell.font = FONT_REMOVED

            # Apply cell-level highlighting for modified fields
            if change_type == "MODIFIED":
                col_name = QUERY6_COLUMNS[col_num - 1]
                # Check if this column corresponds to a changed field
                for snapshot_field, query_col in SNAPSHOT_TO_QUERY6.items():
                    if query_col == col_name and snapshot_field in changed_fields:
                        cell.fill = FILL_CHANGED
                        break

        row_num += 1

    # Auto-fit column widths
    _auto_fit_columns(ws)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    logger.info("Report saved: %s", output_path)
    return output_path


def _write_header(ws):
    """Write and format the header row."""
    for col_num, col_name in enumerate(QUERY6_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _build_row(product_data: dict, snapshot: dict, notes: str) -> list:
    """Build a list of cell values for a query_6 row."""
    agreement_id = snapshot.get("agreement_id", "")
    company = snapshot.get("company", "")
    version = snapshot.get("version", 0)
    amendment_label = "K" if version == 0 else str(version)

    product_name = product_data.get("product_name", "UNKNOWN")
    formulary_status = product_data.get("formulary_status", "")
    frequency = product_data.get("frequency", "Semi-Annual")

    # Derived fields
    company_short = company.split()[0] if company else "UNKNOWN"
    sap_id = f"{company_short}-Comm-{agreement_id}-{agreement_id}"
    freq_short = "S.A" if frequency == "Semi-Annual" else (frequency[:1] if frequency else "S.A")
    program_name = f"{product_name}_{formulary_status}_{freq_short}.O"

    return [
        sap_id,                                                          # SAPID
        amendment_label,                                                 # Amendment #
        agreement_id,                                                    # Contract #
        company,                                                         # Company
        program_name,                                                    # Program Name
        product_name,                                                    # Product
        formulary_status,                                                # Formulary Status / BoB
        product_data.get("base_rebate_pct", ""),                         # Tier Benefit / Base Rebate %
        product_data.get("price_protection_threshold", ""),              # Price Increase Threshold %
        product_data.get("start_date", ""),                              # Program Start Date
        product_data.get("end_date", ""),                                # Program End Date
        snapshot.get("agreement_fields", {}).get("payment_terms_days", ""),  # Payment Terms (Days)
        product_data.get("admin_fee", ""),                               # Admin Fee %
        frequency,                                                       # Frequency
        snapshot.get("commencement_date", ""),                           # Commencement Date
        snapshot.get("termination_date", ""),                            # Termination Date
        snapshot.get("agreement_fields", {}).get("notice_period_days", ""),  # Notice Period (Days)
        product_data.get("minimum_rebate", ""),                          # Minimum Rebate
        product_data.get("maximum_rebate", ""),                          # Maximum Rebate
        product_data.get("market_share", ""),                            # Market Share Requirement
        product_data.get("definition_hash", ""),                         # Definition Hash
        notes,                                                           # Notes
    ]


def _format_agreement_changes(agreement_changes: list) -> str:
    """Format agreement-level changes as a notes string."""
    if not agreement_changes:
        return ""
    parts = []
    for change in agreement_changes:
        field = change["field"].replace("_", " ")
        old = change.get("old", "N/A") or "N/A"
        new = change.get("new", "N/A") or "N/A"
        parts.append(f"Agreement {field}: {old} \u2192 {new}")
    return "; ".join(parts)


def _auto_fit_columns(ws):
    """Auto-fit column widths based on content."""
    for col_num in range(1, len(QUERY6_COLUMNS) + 1):
        col_letter = get_column_letter(col_num)
        max_length = len(QUERY6_COLUMNS[col_num - 1])  # Start with header length
        for row in ws.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        # Cap width and add padding
        adjusted_width = min(max_length + 3, 40)
        ws.column_dimensions[col_letter].width = adjusted_width
